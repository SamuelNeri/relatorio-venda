import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
import logging
import locale
import os
from comissoes import carregar_comissoes

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

# Configuração de log
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class Comissao:
    def __init__(self):
        self.comissoes = carregar_comissoes()

    def calcular_comissao(self, tipo_album, metodo_pagamento, valor_venda):
        if tipo_album not in self.comissoes:
            logging.warning(f"Tipo de álbum desconhecido: {tipo_album}")
            return 0

        if metodo_pagamento not in self.comissoes[tipo_album]:
            logging.warning(f"Método de pagamento desconhecido: {metodo_pagamento}")
            return 0

        porcentagem_comissao = self.comissoes[tipo_album][metodo_pagamento]
        comissao = (porcentagem_comissao / 100) * valor_venda
        return comissao

class SeparadorDeVendas:
    def __init__(self, nome_arquivo):
        self.nome_arquivo = nome_arquivo
        self.df = None
        self.comissao = Comissao()

    def carregar_dados(self):
        self.df = pd.read_excel(self.nome_arquivo)

    def tratar_nans(self):
        self.df.fillna({'Tipo Venda': 'Desconhecido'}, inplace=True)

    def calcular_comissoes(self):
        if self.df is None:
            raise Exception("Dados não carregados. Execute 'carregar_dados()' primeiro.")

        self.tratar_nans()
        self.df['Comissão'] = 0

        colunas_pagamento = ['A FATURAR', 'BOL', 'CC', 'CCRE', 'CD', 'CDEB', 'CTL', 'SGPAYFICTICIO', 'CSGP', 'CHQ',
                             'CHEMP', 'CMBA', 'DAUT', 'DEB', 'DEP', 'DEPONCONTA', 'DIN', 'DOC', 'PAGARME', 'PIX', 'PIXSGP2',
                             'RPAY', 'REP', 'STONEBSBCC', 'STCRED', 'STDEB', 'STONEEMPCC', 'TED', 'TRA']

        for idx, row in self.df.iterrows():
            tipo_album = row['Vendedor']
            for metodo_pagamento in colunas_pagamento:
                if metodo_pagamento in row and pd.notna(row[metodo_pagamento]) and row[metodo_pagamento] > 0:
                    valor_venda = row[metodo_pagamento]
                    comissao = self.comissao.calcular_comissao(tipo_album, metodo_pagamento, valor_venda)
                    self.df.at[idx, 'Comissão'] += comissao

    def separar_por_vendedor(self):
        if self.df is None:
            raise Exception("Dados não carregados. Execute 'carregar_dados()' primeiro.")

        vendas_por_vendedor = self.df.groupby('Vendedor').agg({
            'Valor do Pedido': 'sum',
            'Comissão': 'sum'
        }).reset_index()

        return vendas_por_vendedor

    def gerar_relatorio_detalhado(self):
        if self.df is None:
            raise Exception("Dados não carregados. Execute 'carregar_dados()' primeiro.")

        colunas_pagamento = ['A FATURAR', 'BOL', 'CC', 'CCRE', 'CD', 'CDEB', 'CTL', 'SGPAYFICTICIO', 'CSGP', 'CHQ',
                             'CHEMP', 'CMBA', 'DAUT', 'DEB', 'DEP', 'DEPONCONTA', 'DIN', 'DOC', 'PAGARME', 'PIX', 'PIXSGP2',
                             'RPAY', 'REP', 'STONEBSBCC', 'STCRED', 'STDEB', 'STONEEMPCC', 'TED', 'TRA']

        detalhado = []
        for _, row in self.df.iterrows():
            vendedor = row['Vendedor']
            for metodo in colunas_pagamento:
                if metodo in row and pd.notna(row[metodo]) and row[metodo] > 0:
                    valor = row[metodo]
                    comissao = self.comissao.calcular_comissao(vendedor, metodo, valor)
                    detalhado.append({
                        'Vendedor': vendedor,
                        'Método de Pagamento': metodo,
                        'Valor': valor,
                        'Comissão': comissao
                    })

        df_detalhado = pd.DataFrame(detalhado)
        
        # Calcular totais
        total_valor = df_detalhado['Valor'].sum()
        total_comissao = df_detalhado['Comissão'].sum()
        
        # Adicionar linha de totais
        df_totais = pd.DataFrame([{
            'Vendedor': 'TOTAL',
            'Método de Pagamento': '',
            'Valor': total_valor,
            'Comissão': total_comissao
        }])
        
        df_detalhado = pd.concat([df_detalhado, df_totais], ignore_index=True)

        return df_detalhado

    def gerar_resumo_por_pagamento(self):
        if self.df is None:
            raise Exception("Dados não carregados. Execute 'carregar_dados()' primeiro.")

        colunas_pagamento = ['A FATURAR', 'BOL', 'CC', 'CCRE', 'CD', 'CDEB', 'CTL', 'SGPAYFICTICIO', 'CSGP', 'CHQ',
                             'CHEMP', 'CMBA', 'DAUT', 'DEB', 'DEP', 'DEPONCONTA', 'DIN', 'DOC', 'PAGARME', 'PIX', 'PIXSGP2',
                             'RPAY', 'REP', 'STONEBSBCC', 'STCRED', 'STDEB', 'STONEEMPCC', 'TED', 'TRA']

        resumo = []
        for metodo in colunas_pagamento:
            total_valor = self.df[metodo].sum()
            total_comissao = sum(self.comissao.calcular_comissao(row['Vendedor'], metodo, row[metodo])
                                 for _, row in self.df.iterrows() if pd.notna(row[metodo]) and row[metodo] > 0)
            if total_valor > 0:
                resumo.append({
                    'Método de Pagamento': metodo,
                    'Valor Total': total_valor,
                    'Comissão Total': total_comissao
                })

        df_resumo = pd.DataFrame(resumo)
        
        # Adicionar linha de totais
        total_valor = df_resumo['Valor Total'].sum()
        total_comissao = df_resumo['Comissão Total'].sum()
        df_resumo = pd.concat([df_resumo, pd.DataFrame([{
            'Método de Pagamento': 'TOTAL',
            'Valor Total': total_valor,
            'Comissão Total': total_comissao
        }])], ignore_index=True)

        return df_resumo

    def gerar_relatorio(self, nome_arquivo_saida):
        vendas_por_vendedor = self.separar_por_vendedor()
        detalhado = self.gerar_relatorio_detalhado()
        resumo_pagamento = self.gerar_resumo_por_pagamento()

        with pd.ExcelWriter(nome_arquivo_saida, engine='openpyxl') as writer:
            vendas_por_vendedor.to_excel(writer, sheet_name='Resumo por Vendedor', index=False)
            detalhado.to_excel(writer, sheet_name='Detalhado por Pagamento', index=False)
            resumo_pagamento.to_excel(writer, sheet_name='Resumo por Forma de Pagamento', index=False)

        # Ajustar largura das colunas e formatar
        wb = load_workbook(nome_arquivo_saida)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            self.formatar_planilha(ws)

        wb.save(nome_arquivo_saida)

    def formatar_planilha(self, ws):
        for column in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2

        # Formatar cabeçalhos
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Formatar a última linha (totais)
        last_row = ws.max_row
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=last_row, column=col)
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='double'))

        # Formatar colunas de valores
        for col in range(2, ws.max_column + 1):  # Assumindo que as colunas de valores começam na segunda coluna
            for cell in ws[get_column_letter(col)][2:]:
                cell.number_format = '#,##0.00'

class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Sistema de Relatório de Vendas")
        self.geometry("800x600")
        self.configure(bg='#f0f0f0')

        self.status_var = tk.StringVar()  # Inicializa status_var como atributo da instância
        self.create_widgets()

        self.nome_arquivo = None
        self.separador = None

    def create_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Estilo para os widgets
        style = ttk.Style()
        style.theme_use('clam')

        # Label e Entry para o arquivo selecionado
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill=tk.X, pady=10)

        ttk.Label(file_frame, text="Arquivo selecionado:").pack(side=tk.LEFT)
        self.file_entry = ttk.Entry(file_frame, width=50)
        self.file_entry.pack(side=tk.LEFT, padx=5)

        # Botões
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)

        ttk.Button(button_frame, text="Selecionar Arquivo", command=self.selecionar_arquivo).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Calcular Comissões", command=self.calcular_comissoes).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Mostrar Resultados", command=self.mostrar_resultados).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Salvar Relatório", command=self.salvar_relatorio).pack(side=tk.LEFT, padx=5)

        # Treeview para mostrar resultados
        self.tree = ttk.Treeview(main_frame, columns=('Vendedor', 'Valor das Vendas', 'Comissão'), show='headings')
        self.tree.heading('Vendedor', text='Vendedor')
        self.tree.heading('Valor das Vendas', text='Valor das Vendas')
        self.tree.heading('Comissão', text='Comissão')
        self.tree.pack(fill=tk.BOTH, expand=True, pady=10)

        # Scrollbar para o Treeview
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Barra de status
        self.status_bar = ttk.Label(self, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def selecionar_arquivo(self):
        self.nome_arquivo = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if self.nome_arquivo:
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(0, os.path.basename(self.nome_arquivo))
            self.status_var.set(f"Arquivo selecionado: {os.path.basename(self.nome_arquivo)}")

    def calcular_comissoes(self):
        if not self.nome_arquivo:
            messagebox.showwarning("Nenhum arquivo selecionado", "Por favor, selecione um arquivo primeiro.")
            return

        self.separador = SeparadorDeVendas(self.nome_arquivo)
        self.separador.carregar_dados()
        self.separador.calcular_comissoes()
        self.status_var.set("As comissões foram calculadas com sucesso!")
        self.mostrar_resultados()

    def mostrar_resultados(self):
        if not self.separador:
            messagebox.showwarning("Dados não calculados", "Por favor, calcule as comissões primeiro.")
            return

        vendas_por_vendedor = self.separador.separar_por_vendedor()

        # Limpar a Treeview
        for i in self.tree.get_children():
            self.tree.delete(i)

        # Preencher a Treeview com os resultados
        for _, row in vendas_por_vendedor.iterrows():
            vendedor = row['Vendedor']
            valor_vendas = locale.currency(row['Valor do Pedido'], grouping=True)
            comissao = locale.currency(row['Comissão'], grouping=True)
            self.tree.insert('', 'end', values=(vendedor, valor_vendas, comissao))

        self.status_var.set("Resultados exibidos na tabela.")

    def salvar_relatorio(self):
        if not self.separador:
            messagebox.showwarning("Dados não calculados", "Por favor, calcule as comissões primeiro.")
            return

        nome_arquivo_saida = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if nome_arquivo_saida:
            self.separador.gerar_relatorio(nome_arquivo_saida)
            self.status_var.set(f"Relatório salvo em {nome_arquivo_saida}.")
            messagebox.showinfo("Relatório Salvo", f"O relatório foi salvo em {nome_arquivo_saida} com três planilhas: "
                                "'Resumo por Vendedor', 'Detalhado por Pagamento' e 'Resumo por Forma de Pagamento'.")

if __name__ == "__main__":
    app = App()
    app.mainloop()